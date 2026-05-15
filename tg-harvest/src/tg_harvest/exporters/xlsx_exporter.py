"""Excel (.xlsx) export implementation with formatting."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tg_harvest.exporters.base import BaseExporter, build_row, filter_fields
from tg_harvest.models.parse_result import ParseResult

# Header styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Cell styling
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
TEXT_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

# Column width limits
MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 60
TEXT_COL_WIDTH = 50


class XlsxExporter(BaseExporter):
    async def export(self, result: ParseResult, output_path: Path, file_suffix: str = "") -> Path:
        output_path.mkdir(parents=True, exist_ok=True)

        channel_name = result.channel.username or str(result.channel.id)
        timestamp = result.parsed_at.strftime("%Y%m%d_%H%M%S")
        file_path = output_path / f"{channel_name}_{timestamp}{file_suffix}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"

        # Write header
        for col_idx, field in enumerate(self.fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write data rows
        for row_idx, msg in enumerate(result.messages, 2):
            row = filter_fields(build_row(msg), self.fields)
            for col_idx, field in enumerate(self.fields, 1):
                value = row.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if field == "text":
                    cell.alignment = TEXT_ALIGNMENT
                else:
                    cell.alignment = CELL_ALIGNMENT

        # Auto-fit column widths
        for col_idx, field in enumerate(self.fields, 1):
            if field == "text":
                width = TEXT_COL_WIDTH
            else:
                max_len = len(field)
                for row_idx in range(2, min(len(result.messages) + 2, 102)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        max_len = max(max_len, len(str(cell_value)))
                width = min(max(max_len + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)

            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Add autofilter
        if result.messages:
            last_col = get_column_letter(len(self.fields))
            ws.auto_filter.ref = f"A1:{last_col}{len(result.messages) + 1}"

        # Add channel info sheet
        info_ws = wb.create_sheet("Channel Info")
        info_ws.column_dimensions["A"].width = 20
        info_ws.column_dimensions["B"].width = 40

        info_data = [
            ("Channel ID", result.channel.id),
            ("Title", result.channel.title),
            ("Username", result.channel.username or ""),
            ("Members", result.channel.members_count),
            ("Total Messages", result.total_messages),
            ("Parsed At", result.parsed_at.strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for row_idx, (label, value) in enumerate(info_data, 1):
            label_cell = info_ws.cell(row=row_idx, column=1, value=label)
            label_cell.font = Font(bold=True)
            info_ws.cell(row=row_idx, column=2, value=value)

        wb.save(file_path)
        return file_path
