"""Edge case tests for all exporters: unicode, empty data, special chars."""

import csv
import json
from datetime import datetime, timezone

import pytest

from tg_harvest.exporters.base import build_row, filter_fields, resolve_fields
from tg_harvest.exporters.csv_exporter import CsvExporter
from tg_harvest.exporters.json_exporter import JsonExporter
from tg_harvest.exporters.xlsx_exporter import XlsxExporter
from tg_harvest.models.channel import ChannelInfo
from tg_harvest.models.message import ParsedMessage
from tg_harvest.models.parse_result import ParseResult


def _make_result(messages: list[ParsedMessage], title="Test") -> ParseResult:
    return ParseResult(
        channel=ChannelInfo(id=1, title=title),
        messages=messages,
        parsed_at=datetime(2024, 6, 15, 14, 0, 0, tzinfo=timezone.utc),
    )


def _make_message(msg_id: int, text: str, **kwargs) -> ParsedMessage:
    return ParsedMessage(
        id=msg_id,
        channel_id=1,
        date=datetime(2024, 6, 15, 12, 0, 0, tzinfo=timezone.utc),
        text=text,
        **kwargs,
    )


class TestBuildRowEdgeCases:
    def test_empty_text(self):
        msg = _make_message(1, "")
        row = build_row(msg)
        assert row["text"] == ""

    def test_unicode_text(self):
        msg = _make_message(1, "Привіт світ! 🇺🇦 日本語 العربية")
        row = build_row(msg)
        assert "Привіт" in row["text"]
        assert "🇺🇦" in row["text"]
        assert "日本語" in row["text"]

    def test_newlines_in_text(self):
        msg = _make_message(1, "Line 1\nLine 2\nLine 3")
        row = build_row(msg)
        assert "\n" in row["text"]

    def test_special_chars_in_text(self):
        msg = _make_message(1, 'Quote "test" & <html> tab\there')
        row = build_row(msg)
        assert '"test"' in row["text"]
        assert "&" in row["text"]
        assert "<html>" in row["text"]

    def test_very_long_text(self):
        long_text = "x" * 10000
        msg = _make_message(1, long_text)
        row = build_row(msg)
        assert len(row["text"]) == 10000

    def test_comma_in_text(self):
        msg = _make_message(1, "one, two, three")
        row = build_row(msg)
        assert row["text"] == "one, two, three"

    def test_none_optional_fields(self):
        msg = _make_message(1, "text")
        row = build_row(msg)
        assert row["sender_id"] is None
        assert row["views"] is None
        assert row["media_type"] == ""
        assert row["reactions_total"] == 0


class TestFilterFieldsEdgeCases:
    def test_filter_none_returns_all(self):
        row = {"a": 1, "b": 2, "c": 3}
        assert filter_fields(row, None) == row

    def test_filter_empty_list(self):
        row = {"a": 1, "b": 2}
        result = filter_fields(row, [])
        assert result == {}

    def test_filter_nonexistent_fields_ignored(self):
        row = {"a": 1, "b": 2}
        result = filter_fields(row, ["a", "nonexistent"])
        assert result == {"a": 1}

    def test_filter_preserves_order(self):
        row = {"c": 3, "b": 2, "a": 1}
        result = filter_fields(row, ["c", "a"])
        assert list(result.keys()) == ["c", "a"]


class TestResolveFieldsEdgeCases:
    def test_resolve_none_returns_all(self):
        from tg_harvest.config.constants import ALL_EXPORT_FIELDS

        result = resolve_fields(None)
        assert result == list(ALL_EXPORT_FIELDS)

    def test_resolve_filters_invalid(self):
        result = resolve_fields(["id", "fake_field", "text"])
        assert "id" in result
        assert "text" in result
        assert "fake_field" not in result

    def test_resolve_empty_list(self):
        result = resolve_fields([])
        assert result == []


class TestCsvExporterEdgeCases:
    @pytest.mark.asyncio
    async def test_unicode_messages(self, tmp_path):
        messages = [
            _make_message(1, "Привіт 🇺🇦"),
            _make_message(2, "日本語テスト"),
            _make_message(3, "مرحبا"),
        ]
        result = _make_result(messages)
        path = await CsvExporter().export(result, tmp_path)

        with open(path, encoding="utf-8") as f:
            reader = list(csv.DictReader(f))
        assert len(reader) == 3
        assert reader[0]["text"] == "Привіт 🇺🇦"

    @pytest.mark.asyncio
    async def test_empty_messages(self, tmp_path):
        result = _make_result([])
        path = await CsvExporter().export(result, tmp_path)

        with open(path, encoding="utf-8") as f:
            reader = list(csv.DictReader(f))
        assert len(reader) == 0

    @pytest.mark.asyncio
    async def test_special_chars_in_csv(self, tmp_path):
        messages = [
            _make_message(1, 'He said "hello"'),
            _make_message(2, "value1,value2,value3"),
            _make_message(3, "line1\nline2"),
        ]
        result = _make_result(messages)
        path = await CsvExporter().export(result, tmp_path)

        with open(path, encoding="utf-8") as f:
            reader = list(csv.DictReader(f))
        assert reader[0]["text"] == 'He said "hello"'
        assert reader[1]["text"] == "value1,value2,value3"

    @pytest.mark.asyncio
    async def test_single_field_export(self, tmp_path):
        messages = [_make_message(1, "test")]
        result = _make_result(messages)
        path = await CsvExporter(fields=["id"]).export(result, tmp_path)

        with open(path, encoding="utf-8-sig") as f:
            reader = list(csv.DictReader(f))
        assert list(reader[0].keys()) == ["id"]


class TestJsonExporterEdgeCases:
    @pytest.mark.asyncio
    async def test_unicode_messages(self, tmp_path):
        messages = [
            _make_message(1, "Привіт 🇺🇦"),
            _make_message(2, "日本語テスト"),
        ]
        result = _make_result(messages)
        path = await JsonExporter().export(result, tmp_path)

        data = json.loads(path.read_text(encoding="utf-8"))
        assert data["messages"][0]["text"] == "Привіт 🇺🇦"
        assert data["messages"][1]["text"] == "日本語テスト"

    @pytest.mark.asyncio
    async def test_empty_messages(self, tmp_path):
        result = _make_result([])
        path = await JsonExporter().export(result, tmp_path)

        data = json.loads(path.read_text(encoding="utf-8"))
        assert data["messages"] == []

    @pytest.mark.asyncio
    async def test_field_selection_flattens(self, tmp_path):
        messages = [_make_message(1, "test")]
        result = _make_result(messages)
        path = await JsonExporter(fields=["id", "text"]).export(result, tmp_path)

        data = json.loads(path.read_text(encoding="utf-8"))
        msg = data["messages"][0]
        assert set(msg.keys()) == {"id", "text"}

    @pytest.mark.asyncio
    async def test_special_chars_in_json(self, tmp_path):
        messages = [_make_message(1, 'test "quotes" & <tags>')]
        result = _make_result(messages)
        path = await JsonExporter().export(result, tmp_path)

        data = json.loads(path.read_text(encoding="utf-8"))
        assert data["messages"][0]["text"] == 'test "quotes" & <tags>'


class TestXlsxExporterEdgeCases:
    @pytest.mark.asyncio
    async def test_unicode_messages(self, tmp_path):
        import openpyxl

        messages = [
            _make_message(1, "Привіт 🇺🇦"),
            _make_message(2, "日本語テスト"),
        ]
        result = _make_result(messages)
        path = await XlsxExporter().export(result, tmp_path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Messages"]
        # Find text column
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        text_col = headers.index("text") + 1
        assert ws.cell(row=2, column=text_col).value == "Привіт 🇺🇦"
        assert ws.cell(row=3, column=text_col).value == "日本語テスト"

    @pytest.mark.asyncio
    async def test_empty_messages(self, tmp_path):
        import openpyxl

        result = _make_result([])
        path = await XlsxExporter().export(result, tmp_path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Messages"]
        assert ws.max_row == 1  # Only headers

    @pytest.mark.asyncio
    async def test_single_field_export(self, tmp_path):
        import openpyxl

        messages = [_make_message(1, "test")]
        result = _make_result(messages)
        path = await XlsxExporter(fields=["id"]).export(result, tmp_path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Messages"]
        assert ws.cell(row=1, column=1).value == "id"
        assert ws.max_column == 1

    @pytest.mark.asyncio
    async def test_channel_info_sheet_exists(self, tmp_path):
        import openpyxl

        result = _make_result([_make_message(1, "test")], title="My Channel")
        path = await XlsxExporter().export(result, tmp_path)

        wb = openpyxl.load_workbook(path)
        assert "Channel Info" in wb.sheetnames
        ws = wb["Channel Info"]
        # Check that channel title appears
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "My Channel" in values

    @pytest.mark.asyncio
    async def test_very_long_text(self, tmp_path):
        import openpyxl

        long_text = "x" * 5000
        messages = [_make_message(1, long_text)]
        result = _make_result(messages)
        path = await XlsxExporter().export(result, tmp_path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Messages"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        text_col = headers.index("text") + 1
        assert ws.cell(row=2, column=text_col).value == long_text
