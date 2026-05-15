"""Tests for Excel exporter."""

import pytest
from openpyxl import load_workbook

from tg_harvest.exporters.xlsx_exporter import XlsxExporter


class TestXlsxExporter:
    @pytest.fixture
    def exporter(self):
        return XlsxExporter()

    @pytest.mark.asyncio
    async def test_export_creates_file(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        assert result_path.exists()
        assert result_path.suffix == ".xlsx"

    @pytest.mark.asyncio
    async def test_export_filename_format(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        assert "test_channel" in result_path.name

    @pytest.mark.asyncio
    async def test_export_has_header(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active

        header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        assert "id" in header
        assert "date" in header
        assert "text" in header
        assert "views" in header

    @pytest.mark.asyncio
    async def test_export_row_count(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active
        # 1 header + 2 data rows
        assert ws.max_row == 3

    @pytest.mark.asyncio
    async def test_export_message_data(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active

        # Get header mapping
        header = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

        assert ws.cell(row=2, column=header["id"]).value == 1
        assert (
            ws.cell(row=2, column=header["text"]).value == "Hello world! Check https://example.com"
        )
        assert ws.cell(row=2, column=header["views"]).value == 10000
        assert ws.cell(row=2, column=header["is_pinned"]).value is True

    @pytest.mark.asyncio
    async def test_export_header_styling(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active

        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FFFFFF"
        assert cell.fill.start_color.rgb == "001F4E79"

    @pytest.mark.asyncio
    async def test_export_frozen_panes(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active
        assert ws.freeze_panes == "A2"

    @pytest.mark.asyncio
    async def test_export_autofilter(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active
        assert ws.auto_filter.ref is not None

    @pytest.mark.asyncio
    async def test_export_channel_info_sheet(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)

        assert "Channel Info" in wb.sheetnames
        info_ws = wb["Channel Info"]
        assert info_ws.cell(row=1, column=1).value == "Channel ID"
        assert info_ws.cell(row=1, column=2).value == 1234567890
        assert info_ws.cell(row=2, column=1).value == "Title"
        assert info_ws.cell(row=2, column=2).value == "Test Channel"

    @pytest.mark.asyncio
    async def test_export_creates_directory(self, exporter, sample_parse_result, tmp_path):
        nested = tmp_path / "deep" / "path"
        result_path = await exporter.export(sample_parse_result, nested)
        assert result_path.exists()

    @pytest.mark.asyncio
    async def test_export_with_field_selection(self, sample_parse_result, tmp_path):
        exporter = XlsxExporter(fields=["id", "text", "date"])
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active

        header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        assert header == ["id", "text", "date"]
        assert ws.max_column == 3

    @pytest.mark.asyncio
    async def test_export_minimal_message(self, exporter, sample_parse_result, tmp_path):
        result_path = await exporter.export(sample_parse_result, tmp_path)
        wb = load_workbook(result_path)
        ws = wb.active

        header = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}
        assert ws.cell(row=3, column=header["text"]).value == "Simple message"
        # openpyxl stores empty strings as None
        assert ws.cell(row=3, column=header["media_type"]).value in ("", None)
        assert ws.cell(row=3, column=header["reactions_total"]).value == 0
