"""End-to-end tests for the full export pipeline across all formats."""

import csv
import json

import pytest
from openpyxl import load_workbook

from tg_harvest.config.constants import ALL_EXPORT_FIELDS
from tg_harvest.exporters.csv_exporter import CsvExporter
from tg_harvest.exporters.json_exporter import JsonExporter
from tg_harvest.exporters.xlsx_exporter import XlsxExporter


class TestFullExportPipeline:
    """Test exporting the same data to all formats produces consistent results."""

    @pytest.mark.asyncio
    async def test_all_formats_same_message_count(self, sample_parse_result, tmp_path):
        json_path = await JsonExporter().export(sample_parse_result, tmp_path / "json")
        csv_path = await CsvExporter().export(sample_parse_result, tmp_path / "csv")
        xlsx_path = await XlsxExporter().export(sample_parse_result, tmp_path / "xlsx")

        # JSON
        with open(json_path, encoding="utf-8") as f:
            json_data = json.load(f)
        json_count = len(json_data["messages"])

        # CSV
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            csv_count = sum(1 for _ in csv.DictReader(f))

        # XLSX
        wb = load_workbook(xlsx_path)
        xlsx_count = wb.active.max_row - 1  # minus header

        assert json_count == csv_count == xlsx_count == 2

    @pytest.mark.asyncio
    async def test_all_formats_same_fields_selected(self, sample_parse_result, tmp_path):
        fields = ["id", "text", "date", "views"]

        json_path = await JsonExporter(fields=fields).export(sample_parse_result, tmp_path / "json")
        csv_path = await CsvExporter(fields=fields).export(sample_parse_result, tmp_path / "csv")
        xlsx_path = await XlsxExporter(fields=fields).export(sample_parse_result, tmp_path / "xlsx")

        # JSON
        with open(json_path, encoding="utf-8") as f:
            json_keys = set(json.load(f)["messages"][0].keys())
        assert json_keys == set(fields)

        # CSV
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            csv_keys = set(next(csv.reader(f)))
        assert csv_keys == set(fields)

        # XLSX
        wb = load_workbook(xlsx_path)
        ws = wb.active
        xlsx_keys = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
        assert xlsx_keys == set(fields)

    @pytest.mark.asyncio
    async def test_all_formats_consistent_id_values(self, sample_parse_result, tmp_path):
        """The same message IDs should appear in all formats."""
        json_path = await JsonExporter().export(sample_parse_result, tmp_path / "json")
        csv_path = await CsvExporter().export(sample_parse_result, tmp_path / "csv")
        xlsx_path = await XlsxExporter().export(sample_parse_result, tmp_path / "xlsx")

        # JSON
        with open(json_path, encoding="utf-8") as f:
            json_ids = {m["id"] for m in json.load(f)["messages"]}

        # CSV
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            csv_ids = {int(row["id"]) for row in csv.DictReader(f)}

        # XLSX
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        id_col = header["id"]
        xlsx_ids = {ws.cell(row=r, column=id_col).value for r in range(2, ws.max_row + 1)}

        assert json_ids == csv_ids == xlsx_ids == {1, 2}

    @pytest.mark.asyncio
    async def test_empty_result_all_formats(self, sample_parse_result, tmp_path):
        """Exporting with no messages should produce valid empty files."""
        sample_parse_result.messages = []

        json_path = await JsonExporter().export(sample_parse_result, tmp_path / "json")
        csv_path = await CsvExporter().export(sample_parse_result, tmp_path / "csv")
        xlsx_path = await XlsxExporter().export(sample_parse_result, tmp_path / "xlsx")

        # JSON
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        assert data["messages"] == []
        assert data["total_messages"] == 0

        # CSV — header only
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 1  # header only

        # XLSX — header only
        wb = load_workbook(xlsx_path)
        assert wb.active.max_row == 1

    @pytest.mark.asyncio
    async def test_single_field_export(self, sample_parse_result, tmp_path):
        """Exporting a single field should work for all formats."""
        fields = ["text"]

        json_path = await JsonExporter(fields=fields).export(sample_parse_result, tmp_path / "json")
        csv_path = await CsvExporter(fields=fields).export(sample_parse_result, tmp_path / "csv")
        xlsx_path = await XlsxExporter(fields=fields).export(sample_parse_result, tmp_path / "xlsx")

        # JSON
        with open(json_path, encoding="utf-8") as f:
            msg = json.load(f)["messages"][0]
        assert list(msg.keys()) == ["text"]

        # CSV
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            header = next(csv.reader(f))
        assert header == ["text"]

        # XLSX
        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_column == 1
        assert ws.cell(row=1, column=1).value == "text"

    @pytest.mark.asyncio
    async def test_all_fields_export(self, sample_parse_result, tmp_path):
        """Default (all fields) export should have all columns."""
        csv_path = await CsvExporter().export(sample_parse_result, tmp_path)
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            header = next(csv.reader(f))
        assert header == list(ALL_EXPORT_FIELDS)
